import os
import pandas as pd
from googletrans import Translator
from pypinyin import pinyin, Style
from openpyxl import load_workbook, Workbook
import requests

def get_ipa(word):
    try:
        response = requests.get(f'https://api.dictionaryapi.dev/api/v2/entries/en/{word}')
        response.raise_for_status()
        data = response.json()
        if isinstance(data, list) and 'phonetics' in data[0]:
            phonetics = data[0]['phonetics']
            for phonetic in phonetics:
                if 'text' in phonetic and phonetic['text']:
                    return phonetic['text']
        return ""
    except requests.exceptions.RequestException as e:
        print(f"Error fetching IPA for {word}: {e}")
        return ""

def get_pinyin(chinese_word):
    try:
        pinyin_list = pinyin(chinese_word, style=Style.TONE)
        return ' '.join([item[0] for item in pinyin_list])
    except Exception as e:
        print(f"Error fetching Pinyin for {chinese_word}: {e}")
        return ""

def translate_words(words, output_file_path):
    translator = Translator()

    if os.path.exists(output_file_path):
        workbook = load_workbook(output_file_path)
        # Load or create the 'Crude' sheet
        if 'Crude' in workbook.sheetnames:
            sheet = workbook['Crude']
        else:
            sheet = workbook.create_sheet('Crude')
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Crude'
    
    # Append headers if sheet is empty
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        sheet.append(["English", "IPA", "Vietnamese", "Chinese", "Pinyin"])
    
    existing_data = {row[0].value: row for row in sheet.iter_rows(min_row=2)}

    for idx, word in enumerate(words, 1):
        row_data = existing_data.get(word, None)
        if row_data and all(cell.value for cell in row_data[1:]):
            print(f"........Skipping duplicate word: {word}")
            continue
        
        try:
            print("Loading:...", end="")
            ipa_str = row_data[1].value if row_data and row_data[1].value else get_ipa(word)
            print("...", end="")
            translation_vi = row_data[2].value if row_data and row_data[2].value else translator.translate(word, src='en', dest='vi').text
            print("...", end="")
            translation_zh = row_data[3].value if row_data and row_data[3].value else translator.translate(word, src='en', dest='zh-cn').text
            print("...", end="")
            pinyin_str = row_data[4].value if row_data and row_data[4].value else get_pinyin(translation_zh)
            print("...", end="\n")

            if row_data:
                row_data[1].value = ipa_str if ipa_str else row_data[1].value
                row_data[2].value = translation_vi if translation_vi else row_data[2].value
                row_data[3].value = translation_zh if translation_zh else row_data[3].value
                row_data[4].value = pinyin_str if pinyin_str else row_data[4].value
            else:
                sheet.append([word, ipa_str, translation_vi, translation_zh, pinyin_str])

            print(f"Translated successful word [{idx}]....: {word} -> IPA: {ipa_str}, {translation_vi}, Chinese: {translation_zh}, Pinyin: {pinyin_str}")
        except Exception as e:
            print(f"Error translating {word}........: {str(e)}\n")
            sheet.append([word, "", "", "", ""])

    try:
        workbook.save(output_file_path)
        print(f"File saved successfully at {output_file_path}")
    except Exception as e:
        print(f"Error saving file: {e}")

def main():
    input_path = "access/Second Data/SecondData.csv"
    output_path = "access/Last Data/Data_Learn.xlsx"
    start_line = int(input("Start: "))  # Thay bằng dòng bắt đầu (x)
    end_line = int(input("End: "))  # Thay bằng dòng kết thúc (y)
    
    if not os.path.exists(input_path):
        print(f"Input file not found: {input_path}")
        return

    # Đọc file CSV chỉ các hàng từ start_line đến end_line
    df = pd.read_csv(input_path, skiprows=range(0, start_line), nrows=end_line - start_line, header=None, encoding='utf-8')
    words = df[0].tolist()  # Giả sử dữ liệu cần dịch nằm ở cột 1 (index 0)

    translate_words(words, output_path)

if __name__ == '__main__':
    main()
